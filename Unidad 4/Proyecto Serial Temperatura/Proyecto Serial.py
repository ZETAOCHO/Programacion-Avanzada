# Librerías necesarias
import serial               # Para comunicación serial con el ESP32
import time                 # Para manejo de tiempos de espera
import datetime             # Para obtener la fecha y hora actuales
import os                   # Para verificar si el archivo ya existe
from openpyxl import Workbook, load_workbook           # Para trabajar con archivos .xlsx
from openpyxl.styles import Alignment, Font            # Para centrar texto y poner encabezados en negrita
from openpyxl.utils import get_column_letter           # Para convertir número de columna a letra (por ejemplo: 1 → A)

# Configuración del puerto serial
ser = serial.Serial('COM3', 115200, timeout=1)  # Cambia COM3 si es necesario según tu puerto
time.sleep(10)  # Espera 10 segundos para que el ESP32 termine de iniciar (útil tras el reinicio)

# Envía la hora actual al ESP32 para sincronización
start_time = datetime.datetime.now()  # Captura la fecha y hora actuales
ser.write(start_time.strftime("%Y-%m-%d %H:%M:%S\n").encode())  # Envía en el formato esperado por el ESP32

# Nombre del archivo Excel donde se guardarán los datos
archivo_excel = "temperaturas.xlsx"

# Si el archivo ya existe, lo abre; si no, crea uno nuevo con encabezados
if os.path.exists(archivo_excel):
    wb = load_workbook(archivo_excel)     # Abre archivo existente
    ws = wb.active                        # Selecciona la hoja activa
else:
    wb = Workbook()                       # Crea un nuevo archivo
    ws = wb.active                        # Obtiene la hoja activa
    ws.title = "Lecturas"                # Nombra la hoja
    headers = ["HoraBase", "Temperatura", "Humedad"]  # Encabezados de columna
    ws.append(headers)                   # Agrega encabezados a la primera fila

    # Aplica estilo a los encabezados (negrita y centrado)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

# Bucle principal: lee datos continuamente hasta que se interrumpe manualmente (Ctrl+C)
try:
    while True:
        # Lee una línea desde el puerto serial, ignorando errores de decodificación
        line = ser.readline().decode(errors='ignore').strip()

        # Si la línea tiene contenido y no es un error del sensor
        if line and not line.startswith("Error"):
            print(line)  # Muestra la línea en consola
            parts = line.split(',')  # Divide la línea en partes separadas por comas

            # Si tiene exactamente 3 partes (Hora, Temperatura, Humedad)
            if len(parts) == 3:
                ws.append(parts)  # Agrega la fila al archivo Excel

                # Centra el contenido de la fila recién añadida
                row = ws.max_row  # Obtiene el número de la última fila
                for col in range(1, 4):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

# Este bloque se ejecuta cuando presionas Ctrl+C (KeyboardInterrupt)
except KeyboardInterrupt:
    # Ajusta el ancho de todas las columnas de forma automática
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Obtiene el número de columna (por ejemplo: 1)
        for cell in col:
            if cell.value:  # Si la celda tiene contenido
                max_length = max(max_length, len(str(cell.value)))  # Encuentra el texto más largo
        adjusted_width = max_length + 2  # Ajusta el ancho (+2 para margen)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width  # Aplica el ancho

    # Guarda el archivo Excel con los datos recogidos
    wb.save(archivo_excel)
    print(f"Lecturas guardadas en {archivo_excel}")  # Confirma al usuario
